import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO

# Configuración de estilos para Excel
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_STYLING = True
except ImportError:
    EXCEL_STYLING = False

st.set_page_config(page_title="Gestión de Repuestos Pro", layout="wide")

# --- BANNER SUPERIOR ---
fecha_hoy = datetime.now().strftime("%d/%m/%Y")
st.markdown(f"""
    <style>
    .main-banner {{
        background: linear-gradient(90deg, #1F4E78 0%, #2E75B6 100%);
        padding: 20px;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin-bottom: 20px;
    }}
    </style>
    <div class="main-banner">
        <h1>🛠️ DASHBOARD DE GESTIÓN DE REPUESTOS</h1>
        <p>Reporte Consolidado al <b>{fecha_hoy}</b></p>
    </div>
    """, unsafe_allow_html=True)

uploaded_file = st.file_uploader("Sube el archivo de órdenes", type=["xls", "xlsx", "csv"])

if uploaded_file is not None:
    try:
        # Carga inteligente de datos
        if uploaded_file.name.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(uploaded_file)
        else:
            df = pd.read_csv(uploaded_file, sep=None, engine='python', encoding='latin-1')

        df.columns = df.columns.str.strip()
        
        # --- LÓGICA DE FILTRADO BASE ---
        # Aseguramos que las columnas existan antes de filtrar
        if 'Estado' in df.columns and 'Técnico' in df.columns:
            cond_solicita = df['Estado'].str.contains('Solicita', case=False, na=False)
            es_proceso = df['Estado'].str.contains('Proceso/Repuestos', case=False, na=False)
            tiene_repuestos = df['Repuestos'].astype(str).str.strip().apply(lambda x: len(x) > 2 and x.lower() != 'nan')
            
            patron_excluir = r'^GO|STDIGICENT|STBMDIGI|TCLCUE|TCLCUENC'
            
            mask_estado = cond_solicita | (es_proceso & tiene_repuestos)
            mask_tecnico = ~df['Técnico'].str.upper().str.contains(patron_excluir, na=False, regex=True)
            
            df_base = df[mask_estado & mask_tecnico].copy()
        else:
            st.error("El archivo no contiene las columnas necesarias ('Estado', 'Técnico')")
            st.stop()

        # --- PANEL DE SEGMENTACIÓN (SIDEBAR) ---
        st.sidebar.header("🔍 Segmentación de Datos")
        
        # 1. Segmentar por Técnico
        list_tecnicos = sorted(df_base['Técnico'].dropna().unique())
        tecnicos_sel = st.sidebar.multiselect("Seleccionar Talleres", list_tecnicos, default=list_tecnicos)
        
        # 2. Segmentar por Rango de Fechas
        df_base['Fecha'] = pd.to_datetime(df_base['Fecha'], dayfirst=True, errors='coerce')
        min_date = df_base['Fecha'].min()
        max_date = df_base['Fecha'].max()
        
        if pd.notnull(min_date):
            fecha_sel = st.sidebar.date_input("Rango de Fechas", [min_date, max_date])
        
        # Aplicar Segmentación
        df_filtrado = df_base[df_base['Técnico'].isin(tecnicos_sel)].copy()
        if len(fecha_sel) == 2:
            df_filtrado = df_filtrado[(df_filtrado['Fecha'].dt.date >= fecha_sel[0]) & 
                                      (df_filtrado['Fecha'].dt.date <= fecha_sel[1])]

        if not df_filtrado.empty:
            # MÉTRICAS DINÁMICAS
            m1, m2, m3 = st.columns(3)
            m1.metric("📦 Órdenes en Segmento", len(df_filtrado))
            m2.metric("🧑‍🔧 Talleres Activos", df_filtrado['Técnico'].nunique())
            m3.metric("📅 Fecha más antigua", df_filtrado['Fecha'].min().strftime('%d/%m/%Y'))

            # --- PREPARACIÓN DE COLUMNAS PARA EXCEL ---
            df_filtrado['Enviado'] = "[ ]" 
            columnas_finales = ['Enviado', '#Orden', 'Fecha', 'Técnico', 'Cliente', 'Estado', 'Serie/Artículo', 'Repuestos', 'Producto']
            cols_disp = [c for c in columnas_finales if c in df_filtrado.columns or c == 'Enviado']

            # --- GENERAR EXCEL SEGMENTADO ---
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_filtrado = df_filtrado.sort_values(['Técnico', 'Fecha'])
                df_filtrado[cols_disp].to_excel(writer, index=False, sheet_name='Reporte')
                ws = writer.sheets['Reporte']
                
                if EXCEL_STYLING:
                    # Estilo de cabecera azul oscuro (como tu banner)
                    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = Font(color='FFFFFF', bold=True)
                        cell.alignment = Alignment(horizontal='center')

            st.download_button(
                label="📥 Descargar Segmentación Actual (Excel)",
                data=output.getvalue(),
                file_name=f"Reporte_Segmentado_{fecha_hoy.replace('/','-')}.xlsx",
                use_container_width=True
            )

            # --- VISTA POR PESTAÑAS (SEGMENTACIÓN VISUAL) ---
            st.write("### 📋 Vista Segmentada por Taller")
            tabs = st.tabs([f"📍 {t}" for t in tecnicos_sel if t in df_filtrado['Técnico'].unique()])
            
            for tab, taller in zip(tabs, [t for t in tecnicos_sel if t in df_filtrado['Técnico'].unique()]):
                with tab:
                    df_taller = df_filtrado[df_filtrado['Técnico'] == taller]
                    st.dataframe(df_taller[cols_disp], hide_index=True, use_container_width=True)
                    st.info(f"Total de órdenes para este taller: {len(df_taller)}")
        else:
            st.warning("No hay datos que coincidan con la segmentación elegida.")

    except Exception as e:
        st.error(f"Error en el procesamiento: {e}")
